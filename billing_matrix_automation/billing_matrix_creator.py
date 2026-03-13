from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
from copy import copy
from io import BytesIO

def build_billing_matrix_xlsx(data: dict, template_path: str):
    """
    data is your dict:
      {
        'project_name': ...,
        'project_display_name': ...,
        'month': 'February',
        'rows_data': [ ... ]
      }
    """

    wb = load_workbook(template_path)
    ws = wb.active  # or wb[SHEET_NAME]

    # -------------------------
    # 1) Header / top section
    # -------------------------
    # THESE CELL ADDRESSES ARE EXAMPLES — set them to whatever your template uses.
    # (E.g. you might have Project Manager in B6, Project Name in B10, etc.)
   
    ws["B9"] = data.get("project_name", "")
    ws["B11"] = data.get("project_number", "")  # or project number
    ws["B13"] = ""                   # DGC Requisition Number (if you have it)

    # Month label example: "Req #3 February"
    month = data.get("month", "")
    year = data.get("year", "")
    position = data.get("position", "")
    ws["J13"] = f"{month} {year} Requisition"           

    # -------------------------
    # 2) Table row writing
    # -------------------------
    START_ROW = 24
    rows = data.get("rows_data", [])
    n = len(rows)

    def money(x):
        if x is None or x == "":
            return None
        return float(x)

    def pct(x):
        if not x:
            return None
        s = str(x).strip().replace("%", "")
        return float(s) / 100.0

    # -------------------------
    # Find totals row ("• - Originals")
    # -------------------------
    totals_row = None
    scan_from = START_ROW
    scan_to = START_ROW + 50  # increase if needed

    for r in range(scan_from, scan_to + 1):
        val = ws[f"C{r}"].value
        if isinstance(val, str) and "Originals" in val:
            totals_row = r
            break

    if totals_row is None:
        # If you know it’s ALWAYS row 25 in the template, set totals_row = 25
        raise RuntimeError("Could not find totals row (looked for 'Originals' in column C).")

    # -------------------------
    # Make room: insert rows between START_ROW and totals_row
    # (so totals box is pushed down and never overwritten)
    # -------------------------
    # Template already has 1 formatted data row at START_ROW.
    # If we have n rows, we need (n-1) additional rows.
    rows_to_insert = max(n - 1, 0)

    if rows_to_insert > 0:
        insert_at = START_ROW + 1  # insert right under the prototype row
        ws.insert_rows(insert_at, amount=rows_to_insert)

        # Copy formatting (including row height) from the prototype row to the inserted rows
        prototype_row = START_ROW
        proto_height = ws.row_dimensions[prototype_row].height

        # Copy styles across the full table width: A through R based on your screenshot
        # Adjust end_col if your sheet goes further
        start_col = 1  # A
        end_col = 18   # R

        for rr in range(insert_at, insert_at + rows_to_insert):
            ws.row_dimensions[rr].height = proto_height
            for c in range(start_col, end_col + 1):
                src = ws.cell(row=prototype_row, column=c)
                dst = ws.cell(row=rr, column=c)
                dst._style = copy(src._style)
                dst.number_format = src.number_format
                dst.alignment = copy(src.alignment)
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)

    # After insertion, totals row has moved down automatically.
    # Re-find totals row quickly (same method) because its row index changed:
    totals_row = None
    for r in range(START_ROW, START_ROW + rows_to_insert + 60):
        val = ws[f"C{r}"].value
        if isinstance(val, str) and "Originals" in val:
            totals_row = r
            break
    if totals_row is None:
        raise RuntimeError("Totals row disappeared after insertion (unexpected).")

    # -------------------------
    # Write the data rows
    # -------------------------
    for i, r in enumerate(rows):
        excel_row = START_ROW + i

        ws[f"A{excel_row}"] = r.get("loa_status")
        ws[f"B{excel_row}"] = r.get("comm_no.")
        ws[f"C{excel_row}"] = r.get("subcontractor")
        ws[f"D{excel_row}"] = r.get("trade")
        ws[f"E{excel_row}"] = r.get("cost_code")
        ws[f"F{excel_row}"] = r.get("sub_req_number")
        ws[f"G{excel_row}"] = money(r.get("req_amount"))
        ws[f"H{excel_row}"] = pct(r.get("percent_complete"))
        ws[f"R{excel_row}"] = r.get("notes")

        if ws[f"G{excel_row}"].value is not None:
            ws[f"G{excel_row}"].number_format = '$#,##0.00'
        if ws[f"H{excel_row}"].value is not None:
            ws[f"H{excel_row}"].number_format = '0%'

        ws[f"B{excel_row}"].font = Font(bold=True)

        for col in ["C","D","I","J","K","L"]:
            ws[f"{col}{excel_row}"].alignment = Alignment(wrap_text=True, vertical="center")

        ws[f"I{excel_row}"] = r.get("gl_insurance")
        ws[f"J{excel_row}"] = r.get("auto_insurance")
        ws[f"K{excel_row}"] = r.get("umbrella_insurance")
        ws[f"L{excel_row}"] = r.get("workers_comp_insurance")

    # -------------------------
    # IMPORTANT: Do NOT clear 100 rows anymore.
    # If you want it "tight", you delete only unused rows *above totals*.
    # (This matters when n == 0 or n == 1 or smaller than what template has.)
    # -------------------------

    last_row = START_ROW + n - 1 if n > 0 else START_ROW

    total_cell = f"G{totals_row}"  # <-- assumes the total value cell is in column G on the totals row
    ws[total_cell] = f"=SUM(G{START_ROW}:G{last_row})"
    ws[total_cell].number_format = '$#,##0.00'


    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer